import datetime
import json
import logging
import os
import pprint
import tempfile
import urllib
from io import BytesIO
from zipfile import ZipFile, BadZipFile

import requests
from decouple import config
from django.core.files.storage import default_storage
from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from slack import WebClient
from django.conf import settings
from django.db import IntegrityError

from creative_groups.models import CreativeGroup
from creatives.models import Creative
from background_task import background, exceptions

slack_client = WebClient(config('SLACK_BOT_TOKEN'))

log = logging.getLogger("django")


@background(schedule=1)
def reply_with_instructions(channel):
    slack_client.chat_postMessage(
        channel=channel,
        text="Hello!\n\n"
             "To get started, you\'ll need a template. "
             "If you don\'t have one, respond with _template_\n\n"
             "If you have the template already, fill it out and upload it. "
             "There are columns for:\n\n"
             "*Name* (name each creative something unique)\n\n"
             "and\n\n"
             "*Ad Tag*\n\n"
             "Make sure when copying the ad tags into the "
             "template that you don\'t accidentally add any other characters. "
             "Excel likes to reformat tags sometimes.\n\n"
             "Don\'t worry about whether or not the"
             " tags have blocking. I automatically remove blocking scripts before taking "
             "screenshots. I\'ll upload the results here when I\'m done.\n\n"
             "Shouldn\'t take _too_ long\n"
    )

    log.info('Responded with instructions')


@background(schedule=1)
def reply_with_template(channel):
    slack_client.chat_postMessage(channel=channel, text='Here ya go!')

    slack_client.files_upload(channels=channel, file=os.path.join(settings.MEDIA_ROOT, 'templates/template.xlsx'))

    log.info('Responded with the template')


@background(schedule=1)
def router(request_data, user_name):
    # A function that chooses the correct function depending on the value of A1 in the spreadsheet

    file_info = slack_client.files_info(file=request_data['event']['file_id'])
    file_url = file_info['file']['url_private']

    channel_list = list(file_info['file']['shares']['private'].keys())
    channel = channel_list[0]

    r = requests.get(file_url, headers={'Authorization': 'Bearer %s' % config('SLACK_BOT_TOKEN')})
    r.raise_for_status

    uploaded = default_storage.save(f'uploaded_spreadsheets/{datetime.datetime.now().timestamp()}.xlsx', BytesIO(r.content))

    wb = load_workbook(filename=default_storage.path(uploaded))

    ws = wb.active

    if ws['B1'].value is None:
        campaign_name = 'campaign'
    else:
        campaign_name = ws['B1'].value

    creative_group = CreativeGroup(name=campaign_name)
    creative_group.save()

    for columns in ws.iter_rows(4, ws.max_row, 0, ws.max_column, True):

        try:

            if columns[0] is None or columns[1] is None:  # For blank rows that accidentally get captured as non blank
                log.info('Blank row incorrectly categorized as non-blank -- Skipping')
                continue

            creative = Creative(name=columns[0], markup=columns[1], creative_group_id=creative_group,
                                requested_by=user_name)
            creative.save()

            log.info(f'successfully Created: {creative.name}')

        except IntegrityError:

            log.exception(f'Error Saving Creative: {creative.name}')
            slack_client.chat_postMessage(channel='DSNPWMH88',
                                          text=f"User: {user_name} || "
                                               f"Error: Could not save creative {creative.name}")

        creative.clean_up()
        creative.determine_adserver()
        creative.has_blocking()

        if creative.blocking:
            creative.remove_blocking()

    if ws['A1'].value == 'AdOps Template':
        log.info('Processing for AdOps')
        process_for_ad_ops(creative_group.pk, channel)
    else:
        log.info('Processing for Account Management')
        reply_with_screenshots(creative_group.pk, channel)


@background(schedule=1)
def reply_with_screenshots(creative_group_id, channel):
    creative_group = CreativeGroup.objects.get(pk=creative_group_id)

    log.info(f'Creative Group == {creative_group}')

    progress_meter = slack_client.chat_postMessage(channel=channel, text='◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎')

    errors = []

    p = 1

    log.info(f'Creative Group == {creative_group.creative_set.all()}')

    for creative in creative_group.creative_set.all():

        # Start the progress meter
        meter = progress(p, len(creative_group.creative_set.all()))
        slack_client.chat_update(channel=channel, ts=progress_meter['ts'], text=meter)
        p += 1
        # End the progress meter

        creative.take_screenshot()

        ''' 
        Save_image returns the creative name if there is an error
        push the creative name to an list to return to the user if there are errors
        '''

        creative_name = creative.save_image()

        if creative_name is not None:
            errors.append(creative_name)

    zip_path = f"{settings.MEDIA_ROOT}/zips/{urllib.parse.quote_plus(creative_group.name)}_{datetime.datetime.now().strftime('%m.%d.%H.%M')}.zip"

    with ZipFile(zip_path, 'w', ) as file:

        i = 1

        for creative in creative_group.creative_set.all():

            # If the screenshot image was not saved successfully
            # Do not try to add it to the zip

            try:

                if f'{creative.name}.png' in file.namelist():
                    file.write(creative.screenshot.path, arcname=f'{creative.name}_{i}.png')
                    i += 1

                else:
                    file.write(creative.screenshot.path, arcname=f'{creative.name}.png')

            except ValueError:
                log.error(f'Skipping over {creative.name} because there was no screenshot file')

    slack_client.chat_delete(channel=channel, ts=progress_meter['ts'])  # Delete the progress meter once done

    screenshot_count = creative_group.creative_set.all().count() - len(errors)

    log.info(f'Screenshot count: {screenshot_count}')

    slack_client.chat_postMessage(channel=channel,
                                  text=f'All set. {screenshot_count} screenshots attached below')

    # Only upload the zip file if there are some creatives without errors
    if len(errors) != creative_group.creative_set.all().count():
        slack_client.files_upload(channels=channel, file=zip_path)  # Upload the zip

    if errors:
        slack_client.chat_postMessage(channel=channel,
                                      text=f"I had some issues taking screenshots for the following creatives\n"
                                           f"{errors}\n"
                                           f"Try uploading the template again with the creatives that didnt work."
                                           f"If you receive this message again, "
                                           f"try taking them manually by going to a tag "
                                           f"tester like https://www.cs.iupui.edu/~ajharris/webprog/jsTester.html\n"
                                           f"Sorry for the inconvenience!")


@background(schedule=1)
def reply_with_stats(channel):
    creatives = Creative.objects.all()
    text = f'I have taken {creatives.count()} screenshots'
    slack_client.chat_postMessage(channel=channel, text=text)


# Slash Commands
@background(schedule=1)
def reply_with_preview(text, user, response_url):
    log.info('Replying with preview')
    log.info(f'User: {user} & Response URL: {response_url}')

    creative = Creative()
    creative.name = 'slash command creative'
    creative.requested_by = user
    creative.markup = text
    creative.clean_up()
    creative.determine_adserver()

    if creative.has_blocking():
        creative.remove_blocking()

    creative.take_screenshot()

    creative.save()

    if creative.has_blocking():
        response_string = 'Here is the tag you sent me with blocking removed.'
        markup = creative.markup_without_blocking
    else:
        response_string = 'Here is the tag you sent me'
        markup = creative.markup

    text = f'''
            [
                {{
                    "type": "section",
                    "text": {{
                        "type": "mrkdwn",
                        "text": ":white_check_mark: {response_string}"
                    }}
                }},
                {{
                    "type": "section",
                    "text": {{
                        "type": "mrkdwn",
                        "text": {json.dumps(f'```{markup}```')}
                    }}
                }},
                {{
                    "type": "section",
                    "text": {{
                        "type": "mrkdwn",
                        "text": ":mag: Below is a preview of the creative."
                    }}
                }},
                {{
                    "type": "image",
                    "image_url": "{creative.screenshot_url}",
                    "alt_text": "Ad Tag"
                }},

            ]
            '''

    post_data = {'blocks': text}
    post = requests.post(url=response_url, json=post_data)

    print(text)

    print(post.text)


# Slash Commands
@background(schedule=1)
def reply_with_click_through(text, user, response_url):
    log.info('Replying with click through')
    log.info(f'User: {user} & Response URL: {response_url}')

    creative = Creative()
    creative.name = 'slash command creative'
    creative.requested_by = user
    creative.markup = text
    creative.clean_up()
    creative.determine_adserver()

    if creative.has_blocking():
        creative.remove_blocking()

    creative.validate_click_through()

    creative.save()

    text = f'''
            [
                {{
                    "type": "section",
                    "text": {{
                        "type": "mrkdwn",
                        "text": ":white_check_mark: Click Through Below"
                    }}
                }},
                {{
                    "type": "section",
                    "text": {{
                        "type": "mrkdwn",
                        "text": "{creative.click_through}"
                    }}
                }},
            ]
            '''

    post_data = {'blocks': text}
    post = requests.post(url=response_url, json=post_data)

    print(text)

    print(post.text)


# Helpers
def progress(current_creative, max_creatives):
    status = round(current_creative / max_creatives * 10)

    complete = '◼︎'
    to_do = '◻︎'

    progress_display = f'{complete * status}{to_do * (10 - status)}'

    return progress_display


@background(schedule=1)
def process_for_ad_ops(creative_group_id, channel):
    creative_group = CreativeGroup.objects.get(pk=creative_group_id)

    progress_meter = slack_client.chat_postMessage(channel=channel, text='◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎◻︎')

    errors = []

    p = 1

    for creative in creative_group.creative_set.all():

        # Start the progress meter
        meter = progress(p, len(creative_group.creative_set.all()))
        slack_client.chat_update(channel=channel, ts=progress_meter['ts'], text=meter)
        p += 1
        # End the progress meter

        creative.get_placement_id()
        creative.get_dimensions()
        creative.add_macros()

        creative.take_screenshot()

        creative.validate_click_through()


        ''' 
        Save_image returns the creative name if there is an error
        push the creative name to an list to return to the user if there are errors
        '''

        creative_name = creative.save_image()

        if creative_name is not None:
            errors.append(creative_name)

    # creative_group.validate_click_through()

    out = Workbook()

    dest_filename = default_storage.path(f'saved_spreadsheets/{datetime.datetime.now().timestamp()}.xlsx')

    review = out.active
    review.title = "Review"
    review.column_dimensions['A'].width = 25
    review.column_dimensions['B'].width = 15
    review.column_dimensions['C'].width = 110
    review.column_dimensions['D'].width = 15
    review.column_dimensions['E'].width = 40

    review['A1'] = 'Creative Name'
    review['B1'] = 'Placement ID'
    review['C1'] = 'Mark Up'
    review['D1'] = 'Click Through'
    review['E1'] = 'Preview'

    display = out.create_sheet("Display")
    display.column_dimensions['A'].width = 25
    display.column_dimensions['B'].width = 10
    display.column_dimensions['C'].width = 10
    display.column_dimensions['D'].width = 15
    display.column_dimensions['E'].width = 110
    display.column_dimensions['F'].width = 15
    display.column_dimensions['G'].width = 15

    display['A1'] = 'Creative Name'
    display['B1'] = 'Ad Format'
    display['C1'] = 'Dimensions'
    display['D1'] = 'Clickthrough URL'
    display['E1'] = 'Ad Markup'
    display['F1'] = 'Impression Tracker'
    display['G1'] = 'Placement Name'
    display['H1'] = 'Placement ID'
    display['I1'] = 'MRAID?'
    display['J1'] = 'Markup Type'
    display['K1'] = 'Preview URL'

    row = 2

    for creative in creative_group.creative_set.all():
        log.info(creative.screenshot.path)
        img = Image(creative.screenshot.path)

        review.add_image(img, f'E{row}')

        review.row_dimensions[row].height = creative.height

        name = review.cell(row=row, column=1)
        name.value = creative.name
        name.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        placement_id = review.cell(row=row, column=2)
        placement_id.value = creative.placement_id
        placement_id.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        markup = review.cell(row=row, column=3)
        markup.value = creative.markup
        markup.alignment = Alignment(wrap_text=True, vertical='center')

        click_through = review.cell(row=row, column=4)
        click_through.value = creative.click_through
        click_through.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # DISPLAY SHEET
        display_name = display.cell(row=row, column=1)
        display_name.value = creative.name
        display_name.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        display_ad_format = display.cell(row=row, column=2)
        display_ad_format.value = 'Display'
        display_ad_format.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        display_dimensions = display.cell(row=row, column=3)
        display_dimensions.value = f'{creative.width}x{creative.height}'
        display_dimensions.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        display_click_through = display.cell(row=row, column=4)
        display_click_through.value = creative.click_through
        display_click_through.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        display_ad_markup = display.cell(row=row, column=5)
        display_ad_markup.value = creative.markup_with_macros
        display_ad_markup.alignment = Alignment(wrap_text=True, vertical='center')

        display_pid = display.cell(row=row, column=8)
        display_pid.value = creative.placement_id
        display_pid.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        row += 1

    out.save(filename=dest_filename)

    slack_client.chat_delete(channel=channel, ts=progress_meter['ts'])  # Delete the progress meter once done

    slack_client.files_upload(channels=channel, file=dest_filename)  # Upload the zip
