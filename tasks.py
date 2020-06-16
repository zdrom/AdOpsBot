import pprint

import requests
from decouple import config
from huey.contrib.djhuey import task
from openpyxl import load_workbook
from slack import WebClient
from slack.errors import SlackApiError
from slackeventsapi import SlackEventAdapter
from django.conf import settings

from creative_groups.models import CreativeGroup
from creatives.models import Creative


@task()
def reply_with_instructions(channel):
    slack_client = WebClient(config('SLACK_BOT_TOKEN'))
    slack_client.chat_postMessage(
        channel=channel,
        blocks=[
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": ('Hello! To get started, you\'ll need a template. '
                             'If you don\'t have one, respond with _template_\n'
                             )
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": ('If you have the template already, upload it. '
                             'There is a column for *name* (name each creative something unique), '
                             'and a column for *ad tag*. Make sure when copying the ad tags into the '
                             'template that you don\'t accidentally add any other characters. '
                             'Excel likes to reformat tags sometimes. Don\'t worry about whether or not the'
                             ' tags have blocking. I automatically remove blocking scripts before taking '
                             ' screenshots. I\'ll upload the results here when I\'m done. Shouldn\'t '
                             'take _too_ long\n'
                             )
                }
            }
        ]
    )


@task()
def reply_with_template(channel):
    slack_client = WebClient(config('SLACK_BOT_TOKEN'))
    slack_client.chat_postMessage(channel=channel, text='Here ya go!')
    slack_client.files_upload(channels=channel, file=f'{settings.MEDIA_ROOT}/template.xlsx')


@task()
def reply_with_screenshots(request_data):
    slack_client = WebClient(config('SLACK_BOT_TOKEN'))

    file_info = slack_client.files_info(file=request_data['event']['file_id'])

    channel_list = list(file_info['file']['shares']['private'].keys())
    channel = channel_list[0]

    slack_client.chat_postMessage(channel=channel, text='Confirming receipt. Be back soon.')

    file_url = file_info['file']['url_private']

    r = requests.get(file_url, headers={'Authorization': 'Bearer %s' % config('SLACK_BOT_TOKEN')})
    r.raise_for_status

    with open(f'{settings.MEDIA_ROOT}/uploaded_template.xlsx', 'w+b') as f:
        f.write(bytearray(r.content))

    wb = load_workbook(filename=f'{settings.MEDIA_ROOT}/uploaded_template.xlsx')

    ws = wb.active

    if ws['B1'].value is None:
        campaign_name = 'campaign'
    else:
        campaign_name = ws['B1'].value

    cg = CreativeGroup(name=campaign_name)
    cg.save()

    for columns in ws.iter_rows(4, ws.max_row, 0, ws.max_column, True):
        creative = Creative(name=columns[0], markup=columns[1], creative_group_id=cg)
        creative.save()

        pprint.pprint(f'created {creative.name}')

        creative.determine_adserver()
        creative.has_blocking()

        if creative.blocking:
            creative.remove_blocking()

        creative.take_screenshot()
        creative.save_screenshot()

    file_zip = cg.create_zip()

    slack_client.chat_postMessage(channel=channel, text='Here you go. Screenshots attached below')
    slack_client.files_upload(file=file_zip, channels=channel)